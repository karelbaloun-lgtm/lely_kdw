#!/usr/bin/env python3
"""
Regeneriert den HTML-Bestellformular-Katalog (ITEMS + PEOPLE) direkt aus der
Excel-Datei. Die Excel-Tabelle ist die einzige "Quelle der Wahrheit" fuer
Artikeldaten (Nummer, Bezeichnung, Einheit, Info, Gruppe). Fotos (IMAGES) und
alles andere im HTML (Layout, CSS, JS-Logik) bleiben unveraendert -- Fotos
werden per Artikelnummer aus der bisherigen HTML-Version uebernommen.

Verwendung:
    python3 generate_html.py <excel_datei> <html_datei>

Das Skript ueberschreibt die HTML-Datei in-place (vorher wird eine Kopie
mit ".bak" Endung angelegt).

Erwartete Excel-Struktur (pro Blatt "Grundmaterial" und
"Wuerth Schrauben & Kleinteile"):
  - Zeile 5: Spaltenkoepfe (Artikelnummer | Artikelbezeichnung | Einheit |
    Bestellmenge | zusaetzliche Infos)
  - Ab Zeile 6: entweder eine Gruppen-Kopfzeile (Spalte A leer, Spalte B
    hat Text -> das ist der Gruppenname fuer die folgenden Artikel) oder
    eine Artikelzeile (Spalte A hat die Artikelnummer) oder eine leere
    Trennzeile (beide Spalten leer -> wird ignoriert).
"""
import sys
import re
import json
import openpyxl

SHEET_KATEGORIE = {
    "Grundmaterial": "Grundmaterial",
    "Würth Schrauben & Kleinteile": "Würth Schrauben & Kleinteile",
}


def read_items_from_sheet(ws, kategorie):
    items = []
    current_group = None
    # find header row (row with "Artikelnummer" in col A)
    header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Artikelnummer":
            header_row = r
            break
    if header_row is None:
        raise ValueError(f"Konnte Kopfzeile 'Artikelnummer' nicht finden in Sheet")

    for r in range(header_row + 1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        e = ws.cell(row=r, column=5).value

        a = a.strip() if isinstance(a, str) else a
        b = b.strip() if isinstance(b, str) else b

        if not a and not b:
            continue  # leere Trennzeile
        if not a and b:
            current_group = b
            continue
        if a:
            einheit = (c or "").strip() if isinstance(c, str) else (str(c) if c is not None else "")
            info = (e or "").strip() if isinstance(e, str) else (str(e) if e is not None else "")
            items.append({
                "artikel": str(a),
                "bezeichnung": str(b) if b else "",
                "einheit": einheit,
                "info": info,
                "gruppe": current_group or "",
                "kategorie": kategorie,
                "img": None,
            })
    return items


def read_people_from_sheet(ws):
    people = []
    header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Techniker":
            header_row = r
            break
    if header_row is None:
        raise ValueError("Konnte Kopfzeile 'Techniker' nicht finden in 'Legende Techniker'")
    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        lagerort = ws.cell(row=r, column=2).value
        kostenstelle = ws.cell(row=r, column=3).value
        if not name:
            continue
        people.append({
            "name": str(name).strip(),
            "lagerort": str(lagerort).strip() if lagerort else "",
            "kostenstelle": str(kostenstelle).strip() if kostenstelle is not None else "",
        })
    return people


def extract_js_array(html, var_name):
    marker = f"const {var_name} = ["
    start = html.index(marker)
    array_start = start + len(marker) - 1  # position of "["
    depth = 0
    i = array_start
    in_str = False
    esc = False
    while i < len(html):
        ch = html[i]
        if in_str:
            if esc:
                esc = False
            elif ch == "\\":
                esc = True
            elif ch == '"':
                in_str = False
        else:
            if ch == '"':
                in_str = True
            elif ch == "[":
                depth += 1
            elif ch == "]":
                depth -= 1
                if depth == 0:
                    i += 1
                    break
        i += 1
    array_str = html[array_start:i]
    return start, i + 1, array_str  # start of "const X = ", end after ";", raw array text


def serialize_items(items):
    parts = []
    for it in items:
        img = "null" if it["img"] is None else str(it["img"])
        parts.append(
            '{"artikel": %s, "bezeichnung": %s, "einheit": %s, "info": %s, "gruppe": %s, "kategorie": %s, "img": %s}'
            % (
                json.dumps(it["artikel"], ensure_ascii=False),
                json.dumps(it["bezeichnung"], ensure_ascii=False),
                json.dumps(it["einheit"], ensure_ascii=False),
                json.dumps(it["info"], ensure_ascii=False),
                json.dumps(it["gruppe"], ensure_ascii=False),
                json.dumps(it["kategorie"], ensure_ascii=False),
                img,
            )
        )
    return "[" + ", ".join(parts) + "]"


def serialize_people(people):
    parts = []
    for p in people:
        parts.append(
            '{"name": %s, "lagerort": %s, "kostenstelle": %s}'
            % (
                json.dumps(p["name"], ensure_ascii=False),
                json.dumps(p["lagerort"], ensure_ascii=False),
                json.dumps(p["kostenstelle"], ensure_ascii=False),
            )
        )
    return "[" + ", ".join(parts) + "]"


def main():
    if len(sys.argv) != 3:
        print("Verwendung: python3 generate_html.py <excel_datei> <html_datei>")
        sys.exit(1)

    excel_path, html_path = sys.argv[1], sys.argv[2]

    html = open(html_path, encoding="utf-8").read()

    # --- 1. bisherige ITEMS aus HTML lesen, um Foto-Zuordnung (img) zu erhalten ---
    _, _, old_items_raw = extract_js_array(html, "ITEMS")
    old_items = json.loads(old_items_raw)

    img_by_artikel = {}
    for it in old_items:
        img_by_artikel.setdefault(it["artikel"], []).append(it["img"])

    # --- 2. neue Artikeldaten aus Excel lesen ---
    wb = openpyxl.load_workbook(excel_path, data_only=False)

    new_items = []
    for sheet_name, kategorie in SHEET_KATEGORIE.items():
        ws = wb[sheet_name]
        new_items.extend(read_items_from_sheet(ws, kategorie))

    # Fotos per Artikelnummer uebernehmen (in Reihenfolge, falls Duplikate)
    used_index = {}
    unmatched = []
    for it in new_items:
        art = it["artikel"]
        imgs = img_by_artikel.get(art)
        if imgs:
            idx = used_index.get(art, 0)
            if idx < len(imgs):
                it["img"] = imgs[idx]
                used_index[art] = idx + 1
            else:
                unmatched.append(art)
        else:
            unmatched.append(art)

    # --- 3. PEOPLE aus "Legende Techniker" lesen ---
    people = read_people_from_sheet(wb["Legende Techniker"])

    # --- 4. HTML aktualisieren ---
    start, end, _ = extract_js_array(html, "ITEMS")
    new_items_js = serialize_items(new_items)
    html = html[:start] + f"const ITEMS = {new_items_js};" + html[end:]

    start, end, _ = extract_js_array(html, "PEOPLE")
    new_people_js = serialize_people(people)
    html = html[:start] + f"const PEOPLE = {new_people_js};" + html[end:]

    # --- 5. Backup + Schreiben ---
    with open(html_path + ".bak", "w", encoding="utf-8") as f:
        f.write(open(html_path, encoding="utf-8").read())

    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"Fertig. {len(new_items)} Artikel, {len(people)} Techniker uebernommen.")
    if unmatched:
        print(f"{len(unmatched)} Artikel ohne bestehendes Foto (neu oder Foto-Zuordnung nicht gefunden):")
        for a in unmatched:
            print("  -", a)


if __name__ == "__main__":
    main()
